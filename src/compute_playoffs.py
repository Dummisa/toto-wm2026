"""
TOTO WM 2026 — Playoff-Auswertung.

Liest das Playoff-Workbook (ein Blatt 'Realität' + ein Blatt pro Teilnehmer,
Layout wie die Bracket-Vorlage) und berechnet pro Teilnehmer die Punkte.

Die 32 Startteams im Sechzehntelfinale sind bei allen Blaettern identisch
vorgegeben; ab da werden die Sieger anhand der eingetippten Tore in Python
nachgerechnet (die Excel-Formeln werden NICHT gebraucht).

Punkte (summieren sich ueber die Runden):
  - Richtiger Achtelfinalist  (erreicht Runde der 16):  10 je Team
  - Richtiger Viertelfinalist (erreicht Runde der 8) :  20 je Team
  - Richtiger Halbfinalist    (erreicht Halbfinale)  :  40 je Team
  - Richtiger Finalist        (erreicht Finale)      :  60 je Team
  - Richtiger Weltmeister                            :  80
  - Richtiger 3. Platz                               :  40
  - Pro Spiel mit exakt richtiger Paarung: bis 5 Punkte fuer die Tore
    (−1 je Tor Abweichung, min. 0)
"""

from openpyxl import load_workbook

NON_PLAYER = {'Realität', 'Realitaet', 'POVorlage', 'Vorlage', 'Auswertung'}

# Bracket-Layout (Spalten je Runde) + Zeilen (Teams / Tore)
R32_COLS = [('B', 'C'), ('F', 'G'), ('J', 'K'), ('N', 'O'),
            ('R', 'S'), ('V', 'W'), ('Z', 'AA'), ('AD', 'AE')]
R16_COLS = R32_COLS
QF_COLS = [('D', 'E'), ('L', 'M'), ('T', 'U'), ('AB', 'AC')]
SF_COLS = [('H', 'I'), ('X', 'Y')]
FINAL_COL = ('P', 'Q')
THIRD_COL = ('P', 'Q')

POINTS = {'achtel': 10, 'viertel': 20, 'halb': 40, 'final': 60, 'wm': 80, 'platz3': 40}


def _num(v):
    if v is None or v == '':
        return None
    try:
        return int(v)
    except (TypeError, ValueError):
        try:
            return int(float(v))
        except (TypeError, ValueError):
            return None


def _team(v):
    if v is None:
        return None
    s = str(v).strip()
    return s if s else None


def _winner(home, away, hs, as_):
    if home is None or away is None or hs is None or as_ is None:
        return None
    if hs > as_:
        return home
    if as_ > hs:
        return away
    return None  # K.-o.: Unentschieden = kein Sieger eingetragen


def _loser(home, away, hs, as_):
    w = _winner(home, away, hs, as_)
    if w is None:
        return None
    return away if w == home else home


def read_bracket(ws):
    """Rechnet den kompletten Baum eines Blattes nach.

    Gibt ein dict mit den Mengen je Runde, Champion/Dritter und allen
    Spiel-Slots (fuer den Tor-Bonus) zurueck.
    """
    # ── Sechzehntelfinale (Runde der 32) ──
    # 8 Spiele in Zeile 3 (Tore Zeile 4), 8 in Zeile 8 (Tore Zeile 9)
    r32_top = []   # Sieger der oberen 8 Spiele
    r32_bot = []   # Sieger der unteren 8 Spiele
    slots = []     # (home, away, home_score, away_score) je Spiel-Slot

    for (c1, c2) in R32_COLS:
        h, a = _team(ws[f'{c1}3'].value), _team(ws[f'{c2}3'].value)
        hs, as_ = _num(ws[f'{c1}4'].value), _num(ws[f'{c2}4'].value)
        r32_top.append(_winner(h, a, hs, as_))
        slots.append((h, a, hs, as_))
    for (c1, c2) in R32_COLS:
        h, a = _team(ws[f'{c1}8'].value), _team(ws[f'{c2}8'].value)
        hs, as_ = _num(ws[f'{c1}9'].value), _num(ws[f'{c2}9'].value)
        r32_bot.append(_winner(h, a, hs, as_))
        slots.append((h, a, hs, as_))

    # R16-Teilnehmer: Slot k = (Sieger oberes Spiel k, Sieger unteres Spiel k)
    r16_set = set(t for t in (r32_top + r32_bot) if t)

    # ── Achtelfinale (Runde der 16) ──
    r16_winners = []
    for k, (c1, c2) in enumerate(R16_COLS):
        h, a = r32_top[k], r32_bot[k]
        hs, as_ = _num(ws[f'{c1}14'].value), _num(ws[f'{c2}14'].value)
        r16_winners.append(_winner(h, a, hs, as_))
        slots.append((h, a, hs, as_))
    qf_set = set(t for t in r16_winners if t)

    # ── Viertelfinale (Runde der 8) ──
    qf_winners = []
    for m, (c1, c2) in enumerate(QF_COLS):
        h, a = r16_winners[2 * m], r16_winners[2 * m + 1]
        hs, as_ = _num(ws[f'{c1}19'].value), _num(ws[f'{c2}19'].value)
        qf_winners.append(_winner(h, a, hs, as_))
        slots.append((h, a, hs, as_))
    sf_set = set(t for t in qf_winners if t)

    # ── Halbfinale ──
    sf_winners = []
    sf_losers = []
    for s, (c1, c2) in enumerate(SF_COLS):
        h, a = qf_winners[2 * s], qf_winners[2 * s + 1]
        hs, as_ = _num(ws[f'{c1}24'].value), _num(ws[f'{c2}24'].value)
        sf_winners.append(_winner(h, a, hs, as_))
        sf_losers.append(_loser(h, a, hs, as_))
        slots.append((h, a, hs, as_))
    final_set = set(t for t in sf_winners if t)

    # ── Finale ──
    fc1, fc2 = FINAL_COL
    fh, fa = sf_winners[0], sf_winners[1]
    fhs, fas = _num(ws[f'{fc1}29'].value), _num(ws[f'{fc2}29'].value)
    champion = _winner(fh, fa, fhs, fas)
    slots.append((fh, fa, fhs, fas))

    # ── Spiel um Platz 3 ──
    tc1, tc2 = THIRD_COL
    th, ta = sf_losers[0], sf_losers[1]
    ths, tas = _num(ws[f'{tc1}24'].value), _num(ws[f'{tc2}24'].value)
    third = _winner(th, ta, ths, tas)
    slots.append((th, ta, ths, tas))

    return {
        'r16_set': r16_set,
        'qf_set': qf_set,
        'sf_set': sf_set,
        'final_set': final_set,
        'champion': champion,
        'third': third,
        'slots': slots,
    }


def score_player(player_b, real_b):
    """Punkte eines Spielers gegen die Realität."""
    pts = {
        'achtel': POINTS['achtel'] * len(player_b['r16_set'] & real_b['r16_set']),
        'viertel': POINTS['viertel'] * len(player_b['qf_set'] & real_b['qf_set']),
        'halb': POINTS['halb'] * len(player_b['sf_set'] & real_b['sf_set']),
        'final': POINTS['final'] * len(player_b['final_set'] & real_b['final_set']),
        'wm': POINTS['wm'] if (real_b['champion'] and player_b['champion'] == real_b['champion']) else 0,
        'platz3': POINTS['platz3'] if (real_b['third'] and player_b['third'] == real_b['third']) else 0,
    }

    # Tor-Bonus: pro Slot, wenn Paarung exakt stimmt und beide Tore da sind
    tore = 0
    for (ph, pa, phs, pas), (rh, ra, rhs, ras) in zip(player_b['slots'], real_b['slots']):
        if rh is None or ra is None:
            continue
        if ph == rh and pa == ra:
            if None not in (phs, pas, rhs, ras):
                tore += max(0, 5 - abs(phs - rhs) - abs(pas - ras))
    pts['tore'] = tore

    pts['total'] = sum(pts.values())
    return pts


def compute_playoffs(xlsx_path):
    """Gibt dict zurueck: Spielername -> Punkte-dict (inkl. 'total')."""
    wb = load_workbook(xlsx_path, data_only=False)
    if 'Realität' in wb.sheetnames:
        real = read_bracket(wb['Realität'])
    elif 'Realitaet' in wb.sheetnames:
        real = read_bracket(wb['Realitaet'])
    else:
        raise ValueError("Kein 'Realität'-Sheet im Playoff-Workbook.")

    results = {}
    for name in wb.sheetnames:
        if name in NON_PLAYER:
            continue
        results[name] = score_player(read_bracket(wb[name]), real)
    return results


if __name__ == '__main__':
    import sys
    import json
    path = sys.argv[1] if len(sys.argv) > 1 else 'data/TOTO_Playoffs.xlsx'
    res = compute_playoffs(path)
    for name, p in sorted(res.items(), key=lambda x: -x[1]['total']):
        print(f"{name:14} {p['total']:4}  "
              f"(A{p['achtel']} V{p['viertel']} H{p['halb']} F{p['final']} "
              f"WM{p['wm']} P3{p['platz3']} Tore{p['tore']})")
