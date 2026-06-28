"""
TOTO WM 2026 — Berechnungslogik.

Liest die Roh-Ergebnisse (eingetippte Tore) aus dem Excel-Workbook und
berechnet die gesamte Rangliste neu in Python. Verlaesst sich NICHT auf
die Excel-Formeln (insb. nicht auf das Excel-exklusive SORTBY).

Punktesystem (Gruppenphase):
  - Richtiger Sieger / Unentschieden: 5 Punkte
  - Richtige Toranzahl: max. 5 (minus 1 pro Tor Abweichung, min. 0)
  - Richtiger Sechzehntelfinalist: 5 je korrekt getipptes Team
    -> zaehlt ERST, wenn alle Gruppenspiele in 'Realitaet' eingetragen sind
       UND der Spieler selbst alle 72 Spiele getippt hat.
"""

from openpyxl import load_workbook

COL_PAIRS = [('B', 'C'), ('I', 'J'), ('P', 'Q'),
             ('W', 'X'), ('AD', 'AE'), ('AK', 'AL')]
GRP_TOP = ['A', 'B', 'C', 'D', 'E', 'F']
GRP_BOT = ['G', 'H', 'I', 'J', 'K', 'L']
TOP_ROWS = [(4, 5), (9, 10), (15, 16), (20, 21), (26, 27), (31, 32)]
BOT_ROWS = [(51, 52), (56, 57), (62, 63), (67, 68), (73, 74), (78, 79)]
NON_PLAYER_SHEETS = {'Auswertung', 'Realität', 'Realitaet', 'Anleitung'}


def build_group_layout():
    layout = {}
    for gi, (c1, c2) in enumerate(COL_PAIRS):
        layout[GRP_TOP[gi]] = [(c1, c2, tr, sr) for (tr, sr) in TOP_ROWS]
    for gi, (c1, c2) in enumerate(COL_PAIRS):
        layout[GRP_BOT[gi]] = [(c1, c2, tr, sr) for (tr, sr) in BOT_ROWS]
    return layout


def _num(v):
    if v is None or v == '':
        return None
    try:
        return int(v)
    except (TypeError, ValueError):
        try:
            return int(float(v))
        except (TypeError, ValueError):
            return None


def read_sheet_matches(ws, layout):
    result = {}
    for grp, matches in layout.items():
        ms = []
        for (c1, c2, tr, sr) in matches:
            ht = ws[f'{c1}{tr}'].value
            at = ws[f'{c2}{tr}'].value
            hg = _num(ws[f'{c1}{sr}'].value)
            ag = _num(ws[f'{c2}{sr}'].value)
            ht = ht.strip() if isinstance(ht, str) else ht
            at = at.strip() if isinstance(at, str) else at
            ms.append((ht, at, hg, ag))
        result[grp] = ms
    return result


def compute_standings(matches):
    table = {}

    def ensure(t):
        if t and t not in table:
            table[t] = {'pts': 0, 'gf': 0, 'ga': 0, 'played': 0}

    for ht, at, hg, ag in matches:
        ensure(ht)
        ensure(at)
        if hg is None or ag is None:
            continue
        table[ht]['gf'] += hg
        table[ht]['ga'] += ag
        table[at]['gf'] += ag
        table[at]['ga'] += hg
        table[ht]['played'] += 1
        table[at]['played'] += 1
        if hg > ag:
            table[ht]['pts'] += 3
        elif hg < ag:
            table[at]['pts'] += 3
        else:
            table[ht]['pts'] += 1
            table[at]['pts'] += 1

    ranking = sorted(
        table.keys(),
        key=lambda t: (table[t]['pts'],
                       table[t]['gf'] - table[t]['ga'],
                       table[t]['gf'],
                       t),
        reverse=True)
    return ranking, table


def compute_qualifiers(all_standings):
    qualifiers = set()
    thirds = []
    for grp, (ranking, table) in all_standings.items():
        if len(ranking) >= 1:
            qualifiers.add(ranking[0])
        if len(ranking) >= 2:
            qualifiers.add(ranking[1])
        if len(ranking) >= 3:
            t = ranking[2]
            thirds.append((table[t]['pts'],
                           table[t]['gf'] - table[t]['ga'],
                           table[t]['gf'],
                           t))
    thirds.sort(reverse=True)
    for _, _, _, t in thirds[:8]:
        qualifiers.add(t)
    return qualifiers


def match_points(p_matches, r_matches):
    sieger = 0
    tore = 0
    for (pht, pat, phg, pag), (rht, rat, rhg, rag) in zip(p_matches, r_matches):
        if None in (phg, pag, rhg, rag):
            continue
        ps = (phg > pag) - (phg < pag)
        rs = (rhg > rag) - (rhg < rag)
        if ps == rs:
            sieger += 5
        tore += max(0, 5 - abs(phg - rhg) - abs(pag - rag))
    return sieger, tore


def group_stage_complete(real_matches):
    for grp, ms in real_matches.items():
        for ht, at, hg, ag in ms:
            if hg is None or ag is None:
                return False
    return True


def player_predictions_complete(p_matches):
    for grp, ms in p_matches.items():
        for ht, at, hg, ag in ms:
            if hg is None or ag is None:
                return False
    return True


def compute_playoff_points(player_name, playoff_path=None):
    """Hook fuer spaetere Playoff-Punkte. Aktuell 0."""
    if not playoff_path:
        return 0
    return 0


def find_players(wb):
    return [s for s in wb.sheetnames if s not in NON_PLAYER_SHEETS]


def compute_all(xlsx_path, playoff_path=None):
    wb = load_workbook(xlsx_path, data_only=False)
    layout = build_group_layout()

    if 'Realität' in wb.sheetnames:
        real_ws = wb['Realität']
    elif 'Realitaet' in wb.sheetnames:
        real_ws = wb['Realitaet']
    else:
        raise ValueError("Kein 'Realität'-Sheet gefunden.")

    real_matches = read_sheet_matches(real_ws, layout)
    real_standings = {g: compute_standings(ms) for g, ms in real_matches.items()}
    complete = group_stage_complete(real_matches)
    actual_quali = compute_qualifiers(real_standings) if complete else set()
    playoff_active = bool(playoff_path)

    players = []
    for name in find_players(wb):
        ws = wb[name]
        p_matches = read_sheet_matches(ws, layout)

        sieger_total = 0
        tore_total = 0
        for g in layout:
            s, t = match_points(p_matches[g], real_matches[g])
            sieger_total += s
            tore_total += t

        r32 = 0
        if complete and player_predictions_complete(p_matches):
            p_standings = {g: compute_standings(ms) for g, ms in p_matches.items()}
            p_quali = compute_qualifiers(p_standings)
            r32 = 5 * len(p_quali & actual_quali)

        playoff = compute_playoff_points(name, playoff_path)

        players.append({
            'name': name,
            'sieger': sieger_total,
            'tore': tore_total,
            'r32': r32,
            'playoff': playoff,
            'gruppenphase': sieger_total + tore_total + r32,
            'total': sieger_total + tore_total + r32 + playoff,
        })

    players.sort(key=lambda x: x['total'], reverse=True)
    last_total = None
    last_rank = 0
    for i, p in enumerate(players):
        if p['total'] != last_total:
            last_rank = i + 1
            last_total = p['total']
        p['rank'] = last_rank

    return {
        'group_stage_complete': complete,
        'r32_counts': complete,
        'playoff_active': playoff_active,
        'players': players,
    }
